from openpyxl import *
import os
import random

if __name__ == "__main__":
    print(os.getcwd())
    workbook = load_workbook(filename="Resultats_Concours_ENSAM_PC-SVT-ST.xlsx")
    print(workbook.sheetnames)
    sheet = workbook['Resultat_SE']
    rowCount = sheet.max_row
    random.seed()
    
    for i in range(2, rowCount+1):
        ecoles = [6108, 7102, 5110]
        for letter in ['C', 'D', 'E']:
            k=random.choice(ecoles)
            sheet[letter+str(i)].value = k
            ecoles.remove(k)
            
    workbook.save("Resultats_Concours_ENSAM_PC-SVT-ST.xlsx")