import os
from math import ceil
from os.path import basename
import re
from flask_login.utils import logout_user
from numpy.core.numeric import True_
import pandas as pd
import openpyxl
import base64
import pdfkit
from datetime import date
from zipfile import ZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import Flask,render_template, request, flash, redirect, send_file, send_from_directory, safe_join, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, current_user, login_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from config import UPLOAD_FOLDER, ALLOWED_EXTENSIONS, CHOICE_CODES, SECRET_KEY, ENSAM_CODES, CODES_FILIERES_SM
from models import Results, ResultsSP, LPCasa, LPMeknes, LPRabat, LACasa, LAMeknes, LARabat, LACasaSP, LAMeknesSP, LARabatSP, User, LoginForm, InscritsCasa, InscritsMeknes, InscritsRabat, db
 
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:@localhost/assignementplatform' 
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SECRET_KEY'] = SECRET_KEY
app.config['AVAILABLE_PLACES_SM'] = 'NULL'
app.config['AVAILABLE_PLACES_SP'] = 'NULL'
app.config['LA_NUM'] = 0
login = LoginManager(app)
db.init_app(app)

def allowedFile(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def fileExtension(filename):
    return filename.rsplit('.', 1)[1].lower()

def toPdf(name, typ, listesPrincipales, listesAttentes, AVAILABLE_PLACES_SM, AVAILABLE_PLACES_SP):
    with open('assets/logo'+name.capitalize()+'.png', "rb") as image_file:
        b64Image = base64.b64encode(image_file.read()).decode()
    
    # header = open('assets/'+name+'Header.html', 'r')
    # if not os.path.isdir('html'):
    #     os.mkdir('html')
    # with open('html/'+name+'.html', 'w') as file:
    #     for line in header:
    #         file.write(line.replace('date', str(date.today().year-1)+'-'+str(date.today().year)).replace('$b64Img$', b64Image))
    #     file.write(listesPrincipales[name][['cne','nom']].to_html(index=False, table_id=name+'Table').replace('border="1"', ''))
    #     file.write('</body>')
    # header.close()
    
    if typ == 'LP':
        header = open('assets/'+name+'Header.html', 'r')
        html=''
        for line in header:
            html += line.replace('date', str(date.today().year-1)+'-'+str(date.today().year)).replace('$b64Img$', b64Image)
        html += listesPrincipales[name][['cne','nomPrenom']].to_html(index=False, table_id=name+'Table').replace('border="1"', '')
        html+='</body>'
        header.close()
    elif typ =='LA':
        header = open('assets/'+name+'AttenteHeader.html', 'r')
        html=''
        for line in header:
            html += line.replace('date', str(date.today().year-1)+'-'+str(date.today().year)).replace('$b64Img$', b64Image)
        html += listesAttentes[name][['cne','nomPrenom']].to_html(index=False, table_id=name+'Table').replace('border="1"', '')
        html+='</body>'
        header.close()
    elif typ =='LASP':
        header = open('assets/'+name+'AttenteHeaderSP.html', 'r')
        html=''
        for line in header:
            html += line.replace('date', str(date.today().year-1)+'-'+str(date.today().year)).replace('$b64Img$', b64Image)
        html += listesAttentes[name][['cne','nomPrenom']].to_html(index=False, table_id=name+'Table').replace('border="1"', '')
        html+='</body>'
        header.close()
    options = {
        "enable-local-file-access": None,
        'quiet': ''
    }
    
    if not os.path.isdir('output'):
        os.mkdir('output')
    if typ=='LASP':
        pdfkit.from_string(html, 'output/LA_'+name+'SE.pdf', options=options, css='assets/style.css')
    else:
        pdfkit.from_string(html, 'output/'+typ+'_'+name+'.pdf', options=options, css='assets/style.css')

#
@login.user_loader
def load_user(id):
    return User.query.get(id)

@app.route('/')
def index():
    if current_user.is_authenticated:
        try:
            results = Results.query.order_by(Results.moyenne.desc()).limit(100).all()
            resultsSP = ResultsSP.query.order_by(ResultsSP.moyenne.desc()).limit(100).all()
            lp_casa = LPCasa.query.order_by(LPCasa.nomPrenom).all()
            lp_meknes = LPMeknes.query.order_by(LPMeknes.nomPrenom).all()
            lp_rabat = LPRabat.query.order_by(LPRabat.nomPrenom).all()
            la_casa = LACasa.query.order_by(LACasa.moyenne.desc()).all()
            la_meknes = LAMeknes.query.order_by(LAMeknes.moyenne.desc()).all()
            la_rabat = LARabat.query.order_by(LARabat.moyenne.desc()).all()
            inscritsCasa = InscritsCasa.query.order_by(InscritsCasa.moyenne.desc()).all()
            inscritsMeknes = InscritsMeknes.query.order_by(InscritsMeknes.moyenne.desc()).all()
            inscritsRabat = InscritsRabat.query.order_by(InscritsRabat.moyenne.desc()).all()
        except Exception as e:
            error_text = "<p>The error:<br>" + str(e) + "</p>"
            hed = '<h1>Something is broken.</h1>'
            return hed + error_text
        if not results:
            results = 'NULL'
        elif not resultsSP:
            resultsSP = 'NULL'
        elif not(lp_casa and lp_meknes and lp_rabat):
            lp_casa = 'NULL'
            lp_meknes = 'NULL'
            lp_rabat = 'NULL'
        elif not(la_casa and la_meknes and la_rabat):
            la_casa = 'NULL'
            la_meknes = 'NULL'
            la_rabat = 'NULL'
        elif not inscritsCasa:
            inscritsCasa = 'NULL'
        elif not inscritsMeknes :
            inscritsMeknes = 'NULL'
        elif not inscritsRabat:
            inscritsRabat = 'NULL'
            
        return render_template('index.html', inscritsCasa=inscritsCasa, inscritsMeknes = inscritsMeknes, inscritsRabat=inscritsRabat , results=results, resultsSP=resultsSP, lp_casa=lp_casa, lp_meknes=lp_meknes, lp_rabat=lp_rabat, la_casa=la_casa, la_meknes=la_meknes, la_rabat=la_rabat, AVAILABLE_PLACES_SM=app.config['AVAILABLE_PLACES_SM'], AVAILABLE_PLACES_SP=app.config['AVAILABLE_PLACES_SP'], current_user = current_user)
    else:
        return redirect('/login')

@app.route('/uploadResults', methods=['POST'])
def uploadResults():
    if current_user.is_authenticated:
        if request.method == 'POST':
            if 'results' not in request.files:
                return redirect('/')
            file = request.files['results']
            if file.filename == '':
                return redirect('/')
            if file and allowedFile(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'results.'+fileExtension(filename)))
                if fileExtension(filename) == 'xlsx':
                    resultsDf = pd.read_excel(os.path.join(app.config['UPLOAD_FOLDER'], 'results.xlsx'))
                elif fileExtension(filename) == 'csv':
                    resultsDf = pd.read_csv(os.path.join(app.config['UPLOAD_FOLDER'], 'results.csv'), names=['CNE', 'NOM', 'Choix 1', 'Choix 2', 'Choix 3', 'Moyenne'])
                
                resultsDf = resultsDf[['CNE', 'NOM', 'PRENOM', 'Choix 1', 'Choix 2', 'Choix 3', 'FILIERE', 'CD_FILIERE', 'Moyenne']]
                
                resultsDf = resultsDf.rename(columns={'CNE':'cne', 'NOM':'nom', 'PRENOM':'prenom', 'Choix 1':'choix1', 'CD_FILIERE':'cdFiliere', 'Choix 2':'choix2', 'Choix 3':'choix3', 'FILIERE': 'filiere', 'Moyenne':'moyenne'})
                resultsDf.dropna(subset=['cne', 'nom', 'prenom', 'moyenne', 'choix1', 'choix2', 'choix3', 'cdFiliere'], inplace=True)
                resultsDf['nomPrenom'] = resultsDf['nom'] + ' ' + resultsDf['prenom']
                resultsDf.drop(columns=['nom', 'prenom'], inplace=True)
                re = resultsDf.pop('nomPrenom')
                resultsDf.insert(1, 'nomPrenom', re)
                
                resultsDf['status'] = 0
                resultsSM = resultsDf[resultsDf["cdFiliere"].isin(CODES_FILIERES_SM)]
                resultsSM.sort_values(by=['moyenne'], inplace=True, ascending=False)
                resultsSM.to_sql('results', con=db.engine, index=False, if_exists='replace')

                resultsSP = resultsDf[~resultsDf["cdFiliere"].isin(CODES_FILIERES_SM)]
                resultsSP.sort_values(by=['moyenne'], inplace=True, ascending=False)
                resultsSP.to_sql('results_sp', con=db.engine, index=False, if_exists='replace')
                
                db.session.query(InscritsCasa).delete()
                db.session.query(InscritsMeknes).delete()
                db.session.query(InscritsRabat).delete()
                db.session.query(LPCasa).delete()
                db.session.query(LPMeknes).delete()
                db.session.query(LPRabat).delete()
                db.session.query(LACasa).delete()
                db.session.query(LAMeknes).delete()
                db.session.query(LARabat).delete()
                db.session.commit()
                app.config['LA_NUM'] = 0
                                
                return redirect('/') 

@app.route('/confirmStudents', methods=['POST'])
def confirmStudents():
    if current_user.is_authenticated:
        if request.form['submit'] == 'casa':
            lp = pd.read_sql('SELECT * FROM lp_casa', con=db.engine)
            ResultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
            Results = pd.read_sql('SELECT * FROM results', con=db.engine)
                        
            confirmed = request.form.getlist('confirmedCasa')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('lp_casa', con=db.engine, index=False, if_exists='replace')

            holder=lp[lp['confirmed'] == True]
            for index, row in holder.iterrows():
                if row['choix1'] != ENSAM_CODES['casa']:
                    holder = holder.drop(index)

            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11

            holder=lp[lp['confirmed'] == False]

            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11

            Results.to_sql('results', con=db.engine, index=False, if_exists='replace')
            ResultsSP.to_sql('results_sp', con=db.engine, index=False, if_exists='replace')
            
            RLP = pd.read_sql('SELECT * FROM lp_casa', con=db.engine)
            RLA = pd.read_sql('SELECT * FROM la_casa', con=db.engine)
            ins = pd.read_sql('SELECT * FROM inscritscasa', con=db.engine)
            RLP = RLP.loc[RLP['confirmed'] == True]
            RLA = RLA.loc[RLA['confirmed'] == True]
            
            ins = ins[ins.laNum != 0]
            ins = ins[ins.laNum != app.config['LA_NUM']]
            
            RLP['Ville']='casa'
            RLA['Ville']='casa'
            RLP['laNum']=0
            RLA['laNum']=app.config['LA_NUM']
            
            

            conf = pd.concat([RLP, RLA])
            ins = ins.append(conf, ignore_index=True)
            ins.sort_values(by=['moyenne'], inplace=True, ascending=False)
            
            ins.to_sql('inscritscasa', con=db.engine, index=False, if_exists='replace')
            
        elif request.form['submit'] == 'meknes':
            lp = pd.read_sql('SELECT * FROM lp_meknes', con=db.engine)
            ResultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
            Results = pd.read_sql('SELECT * FROM results', con=db.engine)

            confirmed = request.form.getlist('confirmedMeknes')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('lp_meknes', con=db.engine, index=False, if_exists='replace')

            holder=lp[lp['confirmed'] == True]
            for index, row in holder.iterrows():
                if row['choix1'] != ENSAM_CODES['meknes']:
                    holder = holder.drop(index)

            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11

            holder=lp[lp['confirmed'] == False]

            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11

            Results.to_sql('results', con=db.engine, index=False, if_exists='replace')
            ResultsSP.to_sql('results_sp', con=db.engine, index=False, if_exists='replace')

            
            RLP = pd.read_sql('SELECT * FROM lp_meknes', con=db.engine)
            RLA = pd.read_sql('SELECT * FROM la_meknes', con=db.engine)
            ins = pd.read_sql('SELECT * FROM inscritsmeknes', con=db.engine)
            RLP = RLP.loc[RLP['confirmed'] == True]
            RLA = RLA.loc[RLA['confirmed'] == True]
            
            ins = ins[ins.laNum != 0]
            ins = ins[ins.laNum != app.config['LA_NUM']]
            
            RLP['Ville']='meknes'
            RLA['Ville']='meknes'
            RLP['laNum']=0
            RLA['laNum']=app.config['LA_NUM']  
            

            conf = pd.concat([RLP, RLA])
            ins = ins.append(conf, ignore_index=True)
            ins.sort_values(by=['moyenne'], inplace=True, ascending=False)
            
            ins.to_sql('inscritsmeknes', con=db.engine, index=False, if_exists='replace')
            
        elif request.form['submit'] == 'rabat':
            lp = pd.read_sql('SELECT * FROM lp_rabat', con=db.engine)
            ResultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
            Results = pd.read_sql('SELECT * FROM results', con=db.engine)

            confirmed = request.form.getlist('confirmedRabat')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('lp_rabat', con=db.engine, index=False, if_exists='replace')

            holder=lp[lp['confirmed'] == True]
            for index, row in holder.iterrows():
                if row['choix1'] != ENSAM_CODES['rabat']:
                    holder = holder.drop(index)

            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11


            holder=lp[lp['confirmed'] == False]


            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11

            Results.to_sql('results', con=db.engine, index=False, if_exists='replace')
            ResultsSP.to_sql('results_sp', con=db.engine, index=False, if_exists='replace')

            RLP = pd.read_sql('SELECT * FROM lp_rabat', con=db.engine)
            RLA = pd.read_sql('SELECT * FROM la_rabat', con=db.engine)
            ins = pd.read_sql('SELECT * FROM inscritsrabat', con=db.engine)
            RLP = RLP.loc[RLP['confirmed'] == True]
            RLA = RLA.loc[RLA['confirmed'] == True]
            
            ins = ins[ins.laNum != 0]
            ins = ins[ins.laNum != app.config['LA_NUM']]
            
            RLP['Ville']='rabat'
            RLA['Ville']='rabat'
            RLP['laNum']=0
            RLA['laNum']=app.config['LA_NUM']  
            

            conf = pd.concat([RLP, RLA])
            ins = ins.append(conf, ignore_index=True)
            ins.sort_values(by=['moyenne'], inplace=True, ascending=False)
            
            ins.to_sql('inscritsrabat', con=db.engine, index=False, if_exists='replace')        
    
    return redirect('/')

@app.route('/confirmStudentsLA', methods=['POST'])
def confirmStudentsLA():
    if current_user.is_authenticated:
        if request.form['submit'] == 'casa':
            lp = pd.read_sql('SELECT * FROM la_casa', con=db.engine)
            ResultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
            Results = pd.read_sql('SELECT * FROM results', con=db.engine)

            confirmed = request.form.getlist('confirmedCasa')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('la_casa', con=db.engine, index=False, if_exists='replace')
            ###
            ##
            #
            #do the reallocation
            #remove all confirmed members that were accepted in the LP
            holder=lp[lp['confirmed'] == True]
            for index, row in holder.iterrows():
                if row['choix1'] != ENSAM_CODES['casa']:
                    holder = holder.drop(index)

            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11


            holder=lp[lp['confirmed'] == False]


            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11

            Results.to_sql('results', con=db.engine, index=False, if_exists='replace')
            ResultsSP.to_sql('results_sp', con=db.engine, index=False, if_exists='replace')

            RLP = pd.read_sql('SELECT * FROM lp_casa', con=db.engine)
            RLA = pd.read_sql('SELECT * FROM la_casa', con=db.engine)
            ins = pd.read_sql('SELECT * FROM inscritscasa', con=db.engine)
            RLP = RLP.loc[RLP['confirmed'] == True]
            RLA = RLA.loc[RLA['confirmed'] == True]
            
            ins = ins[ins.laNum != 0]
            ins = ins[ins.laNum != app.config['LA_NUM']]
            
            RLP['Ville']='casa'
            RLA['Ville']='casa'
            RLP['laNum']=0
            RLA['laNum']=app.config['LA_NUM']  
            

            conf = pd.concat([RLP, RLA])
            ins = ins.append(conf, ignore_index=True)
            ins.sort_values(by=['moyenne'], inplace=True, ascending=False)
            
            ins.to_sql('inscritscasa', con=db.engine, index=False, if_exists='replace')


            
        elif request.form['submit'] == 'meknes':
            lp = pd.read_sql('SELECT * FROM la_meknes', con=db.engine)
            ResultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
            Results = pd.read_sql('SELECT * FROM results', con=db.engine)


            confirmed = request.form.getlist('confirmedMeknes')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('la_meknes', con=db.engine, index=False, if_exists='replace')

            holder=lp[lp['confirmed'] == True]
            for index, row in holder.iterrows():
                if row['choix1'] != ENSAM_CODES['meknes']:
                    holder = holder.drop(index)

            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11


            holder=lp[lp['confirmed'] == False]


            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11

            Results.to_sql('results', con=db.engine, index=False, if_exists='replace')
            ResultsSP.to_sql('results_sp', con=db.engine, index=False, if_exists='replace')
            
            RLP = pd.read_sql('SELECT * FROM lp_meknes', con=db.engine)
            RLA = pd.read_sql('SELECT * FROM la_meknes', con=db.engine)
            ins = pd.read_sql('SELECT * FROM inscritsmeknes', con=db.engine)
            RLP = RLP.loc[RLP['confirmed'] == True]
            RLA = RLA.loc[RLA['confirmed'] == True]
            
            ins = ins[ins.laNum != 0]
            ins = ins[ins.laNum != app.config['LA_NUM']]
            
            RLP['Ville']='meknes'
            RLA['Ville']='meknes'
            RLP['laNum']=0
            RLA['laNum']=app.config['LA_NUM']  
            

            conf = pd.concat([RLP, RLA])
            ins = ins.append(conf, ignore_index=True)
            ins.sort_values(by=['moyenne'], inplace=True, ascending=False)
            
            ins.to_sql('inscritsmeknes', con=db.engine, index=False, if_exists='replace')
            
        elif request.form['submit'] == 'rabat':
            lp = pd.read_sql('SELECT * FROM la_rabat', con=db.engine)
            ResultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
            Results = pd.read_sql('SELECT * FROM results', con=db.engine)


            confirmed = request.form.getlist('confirmedRabat')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('la_rabat', con=db.engine, index=False, if_exists='replace')

            holder=lp[lp['confirmed'] == True]
            for index, row in holder.iterrows():
                if row['choix1'] != ENSAM_CODES['rabat']:
                    holder = holder.drop(index)

            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11


            holder=lp[lp['confirmed'] == False]


            Results.loc[Results.cne.isin(holder.cne),'status'] = 11
            ResultsSP.loc[ResultsSP.cne.isin(holder.cne),'status'] = 11

            Results.to_sql('results', con=db.engine, index=False, if_exists='replace')
            ResultsSP.to_sql('results_sp', con=db.engine, index=False, if_exists='replace')
            
            Results.to_sql('results', con=db.engine, index=False, if_exists='replace')
            ResultsSP.to_sql('results_sp', con=db.engine, index=False, if_exists='replace')
            
            RLP = pd.read_sql('SELECT * FROM lp_rabat', con=db.engine)
            RLA = pd.read_sql('SELECT * FROM la_rabat', con=db.engine)
            ins = pd.read_sql('SELECT * FROM inscritsrabat', con=db.engine)
            RLP = RLP.loc[RLP['confirmed'] == True]
            RLA = RLA.loc[RLA['confirmed'] == True]
            
            ins = ins[ins.laNum != 0]
            ins = ins[ins.laNum != app.config['LA_NUM']]
            
            RLP['Ville']='rabat'
            RLA['Ville']='rabat'
            RLP['laNum']=0
            RLA['laNum']=app.config['LA_NUM']            

            conf = pd.concat([RLP, RLA])
            ins = ins.append(conf, ignore_index=True)
            ins.sort_values(by=['moyenne'], inplace=True, ascending=False)
            
            ins.to_sql('inscritsrabat', con=db.engine, index=False, if_exists='replace') 
              
    return redirect('/')

@app.route('/genererLA', methods=['GET', 'POST'])
def genererLA():
#     if current_user.is_authenticated:
#         listesAttentes = {'casa':[], 'meknes':[], 'rabat':[]}
#         listesAttentesSP = {'casa':[], 'meknes':[], 'rabat':[]}
#         listesPrincipales = {'casa':pd.read_sql('SELECT * FROM lp_casa', con=db.engine), 'meknes':pd.read_sql('SELECT * FROM lp_meknes', con=db.engine), 'rabat':pd.read_sql('SELECT * FROM lp_rabat', con=db.engine)}
#         listesReaffectationSM = {}
#         listesReaffectationSP = {}
#         results = pd.read_sql('SELECT * FROM results', con=db.engine)
#         resultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
#         choiceCodes = CHOICE_CODES
#         AVAILABLE_PLACES_SM = {'casa':0, 'meknes':0, 'rabat':0}
#         AVAILABLE_PLACES_SP = {'casa':0, 'meknes':0, 'rabat':0}
        
#         maxAdmisSP = 0
#         maxAdmisSM = 0
#         origin=[]
#         for key in listesPrincipales:
#             listesReaffectationSM[key] = listesPrincipales[key][(listesPrincipales[key]['confirmed']==True) & (listesPrincipales[key]['branche']=='SM')]
#             listesReaffectationSP[key] = listesPrincipales[key][(listesPrincipales[key]['confirmed']==True) & (listesPrincipales[key]['branche']=='SP')]
#             listesReaffectationSM[key]['origin'] = key
#             listesReaffectationSP[key]['origin'] = key
#             maxAdmisSM += len(listesPrincipales[key][listesPrincipales[key]['branche'] == 'SM'].index)
#             maxAdmisSP += len(listesPrincipales[key][listesPrincipales[key]['branche'] == 'SP'].index)
        
#         for key in AVAILABLE_PLACES_SM:
#             AVAILABLE_PLACES_SM[key] = len(listesPrincipales[key][listesPrincipales[key]['branche'] == 'SM'].index)-len(listesReaffectationSM[key].index)
            
#         for key in AVAILABLE_PLACES_SP:
#             AVAILABLE_PLACES_SP[key] = len(listesPrincipales[key][listesPrincipales[key]['branche'] == 'SP'].index)-len(listesReaffectationSP[key].index)
        
#         #do the reallocation
#             #remove all confirmed members that were accepted in the LP
#         for key in listesReaffectationSM:
#             for index, row in listesReaffectationSM[key].iterrows():
#                 if row['choix1'] == ENSAM_CODES[key]:
#                     listesReaffectationSM[key] = listesReaffectationSM[key].drop(index)
#         reaffectedDF = pd.concat([listesReaffectationSM['casa'], listesReaffectationSM['meknes'], listesReaffectationSM['rabat']], ignore_index=True)
#         reaffectedDF = reaffectedDF.sort_values(by=['moyenne'], ascending=False)
        
#         for key in listesReaffectationSP:
#             for index, row in listesReaffectationSP[key].iterrows():
#                 if row['choix1'] == ENSAM_CODES[key]:
#                     listesReaffectationSP[key] = listesReaffectationSP[key].drop(index)
#         reaffectedDFSP = pd.concat([listesReaffectationSP['casa'], listesReaffectationSP['meknes'], listesReaffectationSP['rabat']], ignore_index=True)
#         reaffectedDFSP = reaffectedDFSP.sort_values(by=['moyenne'], ascending=False)
        
#         for index, row in reaffectedDF.iterrows():
#                 if AVAILABLE_PLACES_SM[choiceCodes[row['choix1']]] > 0:
#                     listesAttentes[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
#                     AVAILABLE_PLACES_SM[choiceCodes[row['choix1']]]-=1
#                     AVAILABLE_PLACES_SM[row['origin']]+=1            
#                 elif AVAILABLE_PLACES_SM[choiceCodes[row['choix2']]]>0:            
#                     listesAttentes[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
#                     AVAILABLE_PLACES_SM[choiceCodes[row['choix2']]]-=1   
#                     AVAILABLE_PLACES_SM[row['origin']]+=1            
#                 elif AVAILABLE_PLACES_SM[choiceCodes[row['choix3']]]>0:             
#                     listesAttentes[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
#                     AVAILABLE_PLACES_SM[choiceCodes[row['choix3']]]-=1 
#                     AVAILABLE_PLACES_SM[row['origin']]+=1            
#                 else:
#                     break
        
#         for index, row in reaffectedDFSP.iterrows():
#                 if AVAILABLE_PLACES_SP[choiceCodes[row['choix1']]] > 0:
#                     listesAttentesSP[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
#                     AVAILABLE_PLACES_SP[choiceCodes[row['choix1']]]-=1
#                     AVAILABLE_PLACES_SP[row['origin']]+=1            
#                 elif AVAILABLE_PLACES_SP[choiceCodes[row['choix2']]]>0:            
#                     listesAttentesSP[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
#                     AVAILABLE_PLACES_SP[choiceCodes[row['choix2']]]-=1   
#                     AVAILABLE_PLACES_SP[row['origin']]+=1            
#                 elif AVAILABLE_PLACES_SP[choiceCodes[row['choix3']]]>0:             
#                     listesAttentesSP[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
#                     AVAILABLE_PLACES_SP[choiceCodes[row['choix3']]]-=1 
#                     AVAILABLE_PLACES_SP[row['origin']]+=1            
#                 else:
#                     break
    if current_user.is_authenticated:
        if request.form.get('genererLA') == 'P':
            colNames = ['cne', 'nomPrenom', 'choix1', 'choix2', 'choix3', 'filiere', 'noteMaths', 'notePhysique', 'moyenne']
            PERCENTAGE = {'casa':float(request.form.get('CASA_RANGE')), 'meknes':float(request.form.get('MEKNES_RANGE')), 'rabat':float(request.form.get('RABAT_RANGE'))}
            AVAILABLE_PLACES_SM = {'casa':ceil(int(request.form.get('CASA_MAX_PLACES'))*PERCENTAGE['casa']), 'meknes':ceil(int(request.form.get('MEKNES_MAX_PLACES'))*PERCENTAGE['meknes']), 'rabat':ceil(int(request.form.get('RABAT_MAX_PLACES'))*PERCENTAGE['rabat'])}
            AVAILABLE_PLACES_SP = {'casa':ceil(int(request.form.get('CASA_MAX_PLACES'))*(1-PERCENTAGE['casa'])), 'meknes':ceil(int(request.form.get('MEKNES_MAX_PLACES'))*(1-PERCENTAGE['meknes'])), 'rabat':ceil(int(request.form.get('RABAT_MAX_PLACES'))*(1-PERCENTAGE['rabat']))}            
            listesPrincipales = {'casa':[], 'meknes':[], 'rabat':[]}
            indexes={'casa':0, 'meknes':0, 'rabat':0} 
            choiceCodes=CHOICE_CODES
            
            resultsSM = pd.read_sql('SELECT * FROM results WHERE results.status=0', con=db.engine)
            resultsSP = pd.read_sql('SELECT * FROM results_sp WHERE results_sp.status=0', con=db.engine)
            
            resultsSM.sort_values(by=['moyenne'], inplace=True, ascending=False)
            resultsSM.dropna(subset=['cne', 'nomPrenom', 'moyenne', 'choix1', 'choix2', 'choix3'], inplace=True)
            
            resultsSP.sort_values(by=['moyenne'], inplace=True, ascending=False)
            resultsSP.dropna(subset=['cne', 'nomPrenom', 'moyenne', 'choix1', 'choix2', 'choix3'], inplace=True)
            
            for index, row in resultsSM.iterrows():
                if AVAILABLE_PLACES_SM[choiceCodes[row['choix1']]] > 0:
                    listesPrincipales[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SM[choiceCodes[row['choix1']]]-=1
                    indexes[choiceCodes[row['choix1']]]+=1                
                elif AVAILABLE_PLACES_SM[choiceCodes[row['choix2']]]>0:            
                    listesPrincipales[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SM[choiceCodes[row['choix2']]]-=1
                    indexes[choiceCodes[row['choix2']]]+=1    
                elif AVAILABLE_PLACES_SM[choiceCodes[row['choix3']]]>0:             
                    listesPrincipales[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SM[choiceCodes[row['choix3']]]-=1
                    indexes[choiceCodes[row['choix3']]]+=1  
                else:
                    break
                
            for index, row in resultsSP.iterrows():
                if AVAILABLE_PLACES_SP[choiceCodes[row['choix1']]] > 0:
                    listesPrincipales[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SP[choiceCodes[row['choix1']]]-=1
                    indexes[choiceCodes[row['choix1']]]+=1                
                elif AVAILABLE_PLACES_SP[choiceCodes[row['choix2']]]>0:            
                    listesPrincipales[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SP[choiceCodes[row['choix2']]]-=1
                    indexes[choiceCodes[row['choix2']]]+=1    
                elif AVAILABLE_PLACES_SP[choiceCodes[row['choix3']]]>0:             
                    listesPrincipales[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SP[choiceCodes[row['choix3']]]-=1
                    indexes[choiceCodes[row['choix3']]]+=1  
                else:
                    break
            
            for key in listesPrincipales:
                listesPrincipales[key]=pd.DataFrame(listesPrincipales[key])
                listesPrincipales[key]['confirmed'] = False
                listesPrincipales[key].to_sql('la_'+key, con=db.engine, index=False, if_exists='replace')
            
            app.config['LA_NUM'] += 1
            print(app.config['LA_NUM'])
            return redirect('/')
        elif request.form.get('genererLA') == 'NP':

            colNames = ['cne', 'nomPrenom', 'choix1', 'choix2', 'choix3', 'filiere', 'noteMaths', 'notePhysique', 'moyenne']
            AVAILABLE_PLACES = {'casa':ceil(int(request.form.get('CASA_MAX_PLACES2'))), 'meknes':ceil(int(request.form.get('MEKNES_MAX_PLACES2'))), 'rabat':ceil(int(request.form.get('RABAT_MAX_PLACES2')))}           
            listesPrincipales = {'casa':[], 'meknes':[], 'rabat':[]}
            indexes={'casa':0, 'meknes':0, 'rabat':0} 
            choiceCodes=CHOICE_CODES
            
            resultsSM = pd.read_sql('SELECT * FROM results WHERE results.status=0', con=db.engine)
            resultsSP = pd.read_sql('SELECT * FROM results_sp WHERE results_sp.status=0', con=db.engine)
            
            resultsDf = pd.concat([resultsSM, resultsSP], ignore_index=True)
            
            resultsDf.sort_values(by=['moyenne'], inplace=True, ascending=False)
            resultsDf.dropna(subset=['cne', 'nomPrenom', 'moyenne', 'choix1', 'choix2', 'choix3'], inplace=True)
            nbEtudiants = len(resultsSM.index)+len(resultsSP.index)
            
            for index, row in resultsDf.iterrows():
                if AVAILABLE_PLACES[choiceCodes[row['choix1']]] > 0:
                    listesPrincipales[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES[choiceCodes[row['choix1']]]-=1
                    indexes[choiceCodes[row['choix1']]]+=1                
                elif AVAILABLE_PLACES[choiceCodes[row['choix2']]]>0:            
                    listesPrincipales[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES[choiceCodes[row['choix2']]]-=1
                    indexes[choiceCodes[row['choix2']]]+=1    
                elif AVAILABLE_PLACES[choiceCodes[row['choix3']]]>0:             
                    listesPrincipales[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES[choiceCodes[row['choix3']]]-=1
                    indexes[choiceCodes[row['choix3']]]+=1  
                else:
                    break
            
            for key in listesPrincipales:
                listesPrincipales[key]=pd.DataFrame(listesPrincipales[key])
                listesPrincipales[key]['confirmed'] = False
                listesPrincipales[key].to_sql('la_'+key, con=db.engine, index=False, if_exists='replace')
            
            app.config['LA_NUM'] += 1
            print(app.config['LA_NUM'])
            return redirect('/')          

@app.route('/genererLP', methods=['POST'])
def genererLP():
    if current_user.is_authenticated:
        if request.form.get('genererLP') == 'P':
            colNames = ['cne', 'nomPrenom', 'choix1', 'choix2', 'choix3', 'filiere', 'noteMaths', 'notePhysique', 'moyenne']
            PERCENTAGE = {'casa':float(request.form.get('CASA_RANGE')), 'meknes':float(request.form.get('MEKNES_RANGE')), 'rabat':float(request.form.get('RABAT_RANGE'))}
            AVAILABLE_PLACES_SM = {'casa':ceil(int(request.form.get('CASA_MAX_PLACES'))*PERCENTAGE['casa']), 'meknes':ceil(int(request.form.get('MEKNES_MAX_PLACES'))*PERCENTAGE['meknes']), 'rabat':ceil(int(request.form.get('RABAT_MAX_PLACES'))*PERCENTAGE['rabat'])}
            AVAILABLE_PLACES_SP = {'casa':ceil(int(request.form.get('CASA_MAX_PLACES'))*(1-PERCENTAGE['casa'])), 'meknes':ceil(int(request.form.get('MEKNES_MAX_PLACES'))*(1-PERCENTAGE['meknes'])), 'rabat':ceil(int(request.form.get('RABAT_MAX_PLACES'))*(1-PERCENTAGE['rabat']))}            
            listesPrincipales = {'casa':[], 'meknes':[], 'rabat':[]}
            indexes={'casa':0, 'meknes':0, 'rabat':0} 
            choiceCodes=CHOICE_CODES
            
            resultsSM = pd.read_sql('SELECT * FROM results', con=db.engine)
            resultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
            
            resultsSM.sort_values(by=['moyenne'], inplace=True, ascending=False)
            resultsSM.dropna(subset=['cne', 'nomPrenom', 'moyenne', 'choix1', 'choix2', 'choix3'], inplace=True)
            
            resultsSP.sort_values(by=['moyenne'], inplace=True, ascending=False)
            resultsSP.dropna(subset=['cne', 'nomPrenom', 'moyenne', 'choix1', 'choix2', 'choix3'], inplace=True)
            nbEtudiants = len(resultsSM.index)+len(resultsSP.index)
            
            for index, row in resultsSM.iterrows():
                if AVAILABLE_PLACES_SM[choiceCodes[row['choix1']]] > 0:
                    listesPrincipales[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SM[choiceCodes[row['choix1']]]-=1
                    indexes[choiceCodes[row['choix1']]]+=1                
                elif AVAILABLE_PLACES_SM[choiceCodes[row['choix2']]]>0:            
                    listesPrincipales[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SM[choiceCodes[row['choix2']]]-=1
                    indexes[choiceCodes[row['choix2']]]+=1    
                elif AVAILABLE_PLACES_SM[choiceCodes[row['choix3']]]>0:             
                    listesPrincipales[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SM[choiceCodes[row['choix3']]]-=1
                    indexes[choiceCodes[row['choix3']]]+=1  
                else:
                    break
                
            for index, row in resultsSP.iterrows():
                if AVAILABLE_PLACES_SP[choiceCodes[row['choix1']]] > 0:
                    listesPrincipales[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SP[choiceCodes[row['choix1']]]-=1
                    indexes[choiceCodes[row['choix1']]]+=1                
                elif AVAILABLE_PLACES_SP[choiceCodes[row['choix2']]]>0:            
                    listesPrincipales[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SP[choiceCodes[row['choix2']]]-=1
                    indexes[choiceCodes[row['choix2']]]+=1    
                elif AVAILABLE_PLACES_SP[choiceCodes[row['choix3']]]>0:             
                    listesPrincipales[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES_SP[choiceCodes[row['choix3']]]-=1
                    indexes[choiceCodes[row['choix3']]]+=1  
                else:
                    break
            
            for key in listesPrincipales:
                listesPrincipales[key]=pd.DataFrame(listesPrincipales[key])
                listesPrincipales[key]['confirmed'] = False
                listesPrincipales[key].to_sql('lp_'+key, con=db.engine, index=False, if_exists='replace')
                
            return redirect('/')
        elif request.form.get('genererLP') == 'NP':
            colNames = ['cne', 'nomPrenom', 'choix1', 'choix2', 'choix3', 'filiere', 'noteMaths', 'notePhysique', 'moyenne']
            AVAILABLE_PLACES = {'casa':ceil(int(request.form.get('CASA_MAX_PLACES2'))), 'meknes':ceil(int(request.form.get('MEKNES_MAX_PLACES2'))), 'rabat':ceil(int(request.form.get('RABAT_MAX_PLACES2')))}           
            listesPrincipales = {'casa':[], 'meknes':[], 'rabat':[]}
            indexes={'casa':0, 'meknes':0, 'rabat':0} 
            choiceCodes=CHOICE_CODES
            
            resultsSM = pd.read_sql('SELECT * FROM results', con=db.engine)
            resultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
            
            resultsDf = pd.concat([resultsSM, resultsSP], ignore_index=True)
            
            resultsDf.sort_values(by=['moyenne'], inplace=True, ascending=False)
            resultsDf.dropna(subset=['cne', 'nomPrenom', 'moyenne', 'choix1', 'choix2', 'choix3'], inplace=True)
            nbEtudiants = len(resultsSM.index)+len(resultsSP.index)
            
            for index, row in resultsDf.iterrows():
                if AVAILABLE_PLACES[choiceCodes[row['choix1']]] > 0:
                    listesPrincipales[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES[choiceCodes[row['choix1']]]-=1
                    indexes[choiceCodes[row['choix1']]]+=1                
                elif AVAILABLE_PLACES[choiceCodes[row['choix2']]]>0:            
                    listesPrincipales[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES[choiceCodes[row['choix2']]]-=1
                    indexes[choiceCodes[row['choix2']]]+=1    
                elif AVAILABLE_PLACES[choiceCodes[row['choix3']]]>0:             
                    listesPrincipales[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'moyenne':row['moyenne'], 'cdFiliere': row['cdFiliere']})
                    AVAILABLE_PLACES[choiceCodes[row['choix3']]]-=1
                    indexes[choiceCodes[row['choix3']]]+=1  
                else:
                    break
            
            for key in listesPrincipales:
                listesPrincipales[key]=pd.DataFrame(listesPrincipales[key])
                listesPrincipales[key]['confirmed'] = False
                listesPrincipales[key].to_sql('lp_'+key, con=db.engine, index=False, if_exists='replace')
            
            return redirect('/')

@app.route('/downloadFiles')
def downloadFiles():
    if current_user.is_authenticated:
        listesPrincipales = {'casa':pd.read_sql('SELECT * FROM lp_casa', con=db.engine), 'meknes':pd.read_sql('SELECT * FROM lp_meknes', con=db.engine), 'rabat':pd.read_sql('SELECT * FROM lp_rabat', con=db.engine)}
        listesAttentes = {'casa':pd.read_sql('SELECT * FROM la_casa', con=db.engine), 'meknes':pd.read_sql('SELECT * FROM la_meknes', con=db.engine), 'rabat':pd.read_sql('SELECT * FROM la_rabat', con=db.engine)}
        results = pd.read_sql('SELECT * FROM results', con=db.engine)
        resultsSP = pd.read_sql('SELECT * FROM results_sp', con=db.engine)
        
        nbEtudiants = len(results.index)
        nbEtudiantsSP = len(resultsSP.index)
        
        wb = openpyxl.load_workbook('output/Resultats.xlsx')
        ws = wb['resultatsSM']
        ws.delete_rows(0,4*nbEtudiants)
        for row in dataframe_to_rows(results.loc[:,~(results.columns.str.match("Unnamed")) & ~(results.columns.str.match("status"))], index=False):
            ws.append(row)
        maxCol = ws.max_column
        ws.cell(1, maxCol+1).value = 'Affectation Liste Principale'
        ws.cell(1, maxCol+2).value = 'Affectation Liste d\'Attente'
        
        for i in range(2,nbEtudiants+2):
            flag=1
            for key in listesPrincipales:
                if ws['A'+str(i)].value in listesPrincipales[key].cne.values:
                    ws.cell(i, maxCol+1).value = key.capitalize()
                    flag=0
            if flag:
                ws.cell(i, maxCol+1).value = ''
        for i in range(2,nbEtudiants+2):
            flag=1
            for key in listesAttentes:
                if ws['A'+str(i)].value in listesAttentes[key].cne.values:
                    ws.cell(i, maxCol+2).value = key.capitalize()
                    flag = 0
            if flag:
                ws.cell(i, maxCol+2).value = ''

        ws = wb['resultatsSP']
        ws.delete_rows(0,4*nbEtudiantsSP)
        for row in dataframe_to_rows(resultsSP.loc[:,~resultsSP.columns.str.match("Unnamed") & ~(results.columns.str.match("status"))], index=False):
            ws.append(row)
        maxColSP = ws.max_column
        ws.cell(1, maxColSP+1).value = 'Affectation Liste Principale'
        ws.cell(1, maxColSP+2).value = 'Affectation Liste d\'Attente'
        for i in range(2,nbEtudiantsSP+2):
            flag=1
            for key in listesPrincipales:
                if ws['A'+str(i)].value in listesPrincipales[key].cne.values:
                    ws.cell(i, maxColSP+1).value = key.capitalize()
                    flag=0
            if flag:
                ws.cell(i, maxColSP+1).value = ''
        for i in range(2,nbEtudiantsSP+2):
            flag=1
            for key in listesAttentes:
                if ws['A'+str(i)].value in listesAttentes[key].cne.values:
                    ws.cell(i, maxColSP+2).value = key.capitalize()
                    flag = 0
            if flag:
                ws.cell(i, maxColSP+2).value = ''
    
    
        for key in listesPrincipales:               
            name = 'LP_'+key
            if name not in wb.sheetnames:
                wb.create_sheet(name)
                ws = wb[name]
            else:
                ws = wb[name]
                ws.delete_rows(2,2*nbEtudiants)
            
            for row in dataframe_to_rows(listesPrincipales[key].drop('confirmed', axis=1), index=False):
                ws.append(row)
                        
            listesPrincipales[key].sort_values(by=['nomPrenom'], inplace=True)
            toPdf(key, 'LP', listesPrincipales, listesAttentes, app.config['AVAILABLE_PLACES_SM'], app.config['AVAILABLE_PLACES_SP'])
        for key in listesAttentes:               
            name = 'LA_'+key
            if name not in wb.sheetnames:
                wb.create_sheet(name)
                ws = wb[name]
            else:
                ws = wb[name]
                ws.delete_rows(2,2*nbEtudiants)
            
            for row in dataframe_to_rows(listesAttentes[key], index=False):
                ws.append(row)
            
            toPdf(key, 'LA', listesPrincipales, listesAttentes, app.config['AVAILABLE_PLACES_SM'], app.config['AVAILABLE_PLACES_SP'])        
        
        ws = wb['resultatsSM']
        for cell in ws["1:1"]:
            cell.font = openpyxl.styles.Font(color='00000000', bold=True, size='12') 
            cell.alignment = openpyxl.styles.alignment.Alignment(horizontal = 'center', vertical ='center')
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), right=openpyxl.styles.borders.Side(style='thin'), top=openpyxl.styles.borders.Side(style='thin'), bottom=openpyxl.styles.borders.Side(style='thin'))
        
        ws = wb['resultatsSP']
        for cell in ws["1:1"]:
            cell.font = openpyxl.styles.Font(color='00000000', bold=True, size='12') 
            cell.alignment = openpyxl.styles.alignment.Alignment(horizontal = 'center', vertical ='center')
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), right=openpyxl.styles.borders.Side(style='thin'), top=openpyxl.styles.borders.Side(style='thin'), bottom=openpyxl.styles.borders.Side(style='thin')) 
        
        wb.save('output/Resultats.xlsx')
        return redirect('/zipnsend')


@app.route('/downloadInscrits', methods=['POST'])
def downloadInscrits():
    InscritsCasa = pd.read_sql('SELECT * FROM inscritscasa', con=db.engine)
    InscritsMeknes = pd.read_sql('SELECT * FROM inscritsmeknes', con=db.engine)
    InscritsRabat = pd.read_sql('SELECT * FROM inscritsrabat', con=db.engine)
    
    nbEtudiantsC = len(InscritsCasa.index)
    nbEtudiantsM = len(InscritsMeknes.index)
    nbEtudiantsR = len(InscritsRabat.index)
    
    wb = openpyxl.load_workbook('output/Inscrits.xlsx')
    ws = wb['ConfirmesCasa']
    ws.delete_rows(0,4*nbEtudiantsC)
    for row in dataframe_to_rows(InscritsCasa.loc[:,~(InscritsCasa.columns.str.match("Unnamed")) & ~(InscritsCasa.columns.str.match("confirmed")) & ~(InscritsCasa.columns.str.match("Ville")) & ~(InscritsCasa.columns.str.match("laNum"))], index=False):
        ws.append(row)
    for cell in ws["1:1"]:
            cell.font = openpyxl.styles.Font(color='00000000', bold=True, size='12') 
            cell.alignment = openpyxl.styles.alignment.Alignment(horizontal = 'center', vertical ='center')

    ws = wb['ConfirmesMeknes']
    ws.delete_rows(0,4*nbEtudiantsM)
    for row in dataframe_to_rows(InscritsMeknes.loc[:,~(InscritsMeknes.columns.str.match("Unnamed")) & ~(InscritsMeknes.columns.str.match("confirmed")) & ~(InscritsMeknes.columns.str.match("Ville")) & ~(InscritsMeknes.columns.str.match("laNum"))], index=False):
        ws.append(row)
    for cell in ws["1:1"]:
            cell.font = openpyxl.styles.Font(color='00000000', bold=True, size='12') 
            cell.alignment = openpyxl.styles.alignment.Alignment(horizontal = 'center', vertical ='center')
     
    ws = wb['ConfirmesRabat']
    ws.delete_rows(0,4*nbEtudiantsR)
    for row in dataframe_to_rows(InscritsRabat.loc[:,~(InscritsRabat.columns.str.match("Unnamed")) & ~(InscritsRabat.columns.str.match("confirmed")) & ~(InscritsRabat.columns.str.match("Ville")) & ~(InscritsRabat.columns.str.match("laNum"))], index=False):
        ws.append(row)
    for cell in ws["1:1"]:
            cell.font = openpyxl.styles.Font(color='00000000', bold=True, size='12') 
            cell.alignment = openpyxl.styles.alignment.Alignment(horizontal = 'center', vertical ='center')
    
    wb.save('output/Inscrits.xlsx')
    return redirect('/output/Inscrits.xlsx')

@app.route('/output/Inscrits.xlsx')
def getInscrits():
    if current_user.is_authenticated:
        try:
            return send_from_directory('output', 'Inscrits.xlsx', as_attachment=True)
        except FileNotFoundError:
            abort(404) 

@app.route('/zipnsend')
def zipnsend():
    if current_user.is_authenticated:
        zipObj = ZipFile('zip/output.zip', 'w')
        for folderName, subfolders, filenames in os.walk('output'):
            for filename in filenames:
                #create complete filepath of file in directory
                filePath = os.path.join(folderName, filename)
                # Add file to zip
                zipObj.write(filePath, basename(filePath))
        
        return redirect('/zip/output.zip')

@app.route("/zip/output.zip")
def getFile():
    if current_user.is_authenticated:
        try:
            return send_from_directory('zip', 'output.zip', as_attachment=True)
        except FileNotFoundError:
            abort(404)           

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect('/')
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid username or password')
            return redirect('/login')
        login_user(user)
        return redirect('/')
    return render_template('login.html', title='Sign In', form=form)

@app.route('/logout')
def logout():
    if current_user.is_authenticated:
        logout_user()
    return redirect('/')

if __name__ == '__main__':
    app.run(host='localhost', port=5000, debug=True)